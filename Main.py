from PyQt4 import QtCore, QtGui, Qt
import threading
from PyQt4.QtGui import *
import os
from os import path
import sys
import time
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
#from datetime import date, timedelta
import sqlite3
import pandas as pd
import csv
import numpy as np
import matplotlib.pyplot as plt
import getpass
import ConfigParser
Config = ConfigParser.ConfigParser()

try:
    _fromUtf8 = QtCore.QString.fromUtf8
except AttributeError:
    def _fromUtf8(s):
        return s
try:
    _encoding = QtGui.QApplication.UnicodeUTF8

    def _translate(context, text, disambig):
        return QtGui.QApplication.translate(context, text, disambig, _encoding)
except AttributeError:
    def _translate(context, text, disambig):
        return QtGui.QApplication.translate(context, text, disambig)


class MainWindow(QtGui.QMainWindow):
    def __init__(self, parent=None):
        QtGui.QWidget.__init__(self, parent)
        # Todo: Take this out and make more dynamic
        # self.ConnectODBC()
        super(MainWindow, self).__init__()
        # outside = Outside()
        # self.connect(outside, QtCore.SIGNAL("progress(int, int)"), self.progress)
        # self.connect(QMainWindow.SIGNAL("progress(int, int)"), self.progress)
        self.setGeometry(1930, 30, 1000, 1013)  # Make dynamic
        self.setWindowTitle("Scripts")
        # self.setWindowIcon(QtGui.QIcon('eye.png'))
        extractAction = QtGui.QAction("&Quit", self)
        extractAction.setShortcut("Ctrl+Q")
        extractAction.setStatusTip('Leave The App')
        extractAction.triggered.connect(self.close_application)
        # Add StatusBar to Main Window
        # self.statusBar()
        # Add MenuBar to Main Window (and add menus)
        '''
        mainMenu = self.menuBar()
        fileMenu = mainMenu.addMenu('&File')
        fileMenu.addAction(extractAction)
        editMenu = mainMenu.addMenu('&Edit')
        editMenu.addAction(extractAction)
        '''
        # Add toolbar
        self.home()
        # self.progress(0,4000)
        # Start the Main User-interface:
        self.mainUi(self)

    def writeinifile(self):
        # GetMonitorsize()
        projiniFileName = 'projini.txt'
        projiniFileNameOld = 'projini_old.txt'
        now = time.time()
        userName = getpass.getuser()
        currentDirectory = os.getcwd()
        currentDateTime = time.time()

    def checkinifile(self):
        config = ConfigParser.ConfigParser()
        #GetMonitorsize()
        projiniFileName = 'proj.ini'
        projiniFileNameOld = 'projbak.ini'
        now = time.time()
        userName = getpass.getuser()
        currentDirectory = os.getcwd()
        currentDateTime = time.time()
        #pnumber = os.getpid()
        #fullFilePathName = (currentDirectory + '/' + projiniFileName)
        #fullFilePathNameOld = (currentDirectory + '/' + projiniFileNameOld)
        #listofNetifaces = netifaces.interfaces()
        #countofNetifaces = ('{}'.format(len(listofNetifaces)))

        #fileobj = open(projiniFileName, 'w')

        if not (path.isfile(projiniFileName)):

            #fileobj.write(currentDirectory + '\n')
            #fileobj.write('{}'.format(now) + '\n')
            #fileobj.write(userName + '\n')
            #fileobj.write("{}".format(pnumber) + '\n')
            #fileobj.write("{}".format(nmons) + '\n')
            #for m in range(nmons):
            #    mg = screen.get_monitor_geometry(m)
            #    msize = "%d,%d,%d" % (m, mg.width, mg.height)
            #    monitors.append(msize)
            #fileobj.write("{}".format(monitors) + '\n')
            config.add_section('MainConfig')
            config.set('MainConfig', 'initialDateTime', currentDateTime)
            #fileobj.write("{}".format('0') + '\n')
            #fileobj.write("{}".format('0') + '\n')
            #fileobj.write("{}".format(currentDateTime) + '\n')
            #with open(path, projiniFileName) as config_file:
            with open(projiniFileName, "wb") as config_file:
                config.write(config_file)

            #fileobj.close()
        else:
            #if (path.isfile(fullFilePathNameOld)):
            #    os.remove(projiniFileNameOld)
            #else:
            #    pass
            config.read(projiniFileName)
            #generalConfig = config.get(config.sections('MainConfig'))
            #countOfBankAccs = (len(config.get(config.sections())-1))
            lastTime = config.get('MainConfig','initialDateTime')
            print ("File Created: {}".format(lastTime))
            print ("TimeDelta: {}".format(now-lastTime))
            #fileobj.close()

            #os.rename(projiniFileName, projiniFileNameOld)
            #fileObjOld = open(projiniFileNameOld, 'r')
            #location = fileObjOld.readline()
            #lastOpened = fileObjOld.readline()
            #lastUserName = fileObjOld.readline()
            #lastpnumber = fileObjOld.readline()
            #lastQtyMons = fileObjOld.readline()
            #lastMonitorsizes = fileObjOld.readline()
            #lastUsedMonitor = fileObjOld.readline()
            #lastX = fileObjOld.readline()
            #lastY = fileObjOld.readline()
            #fileObjOld.close()
            #fileobj = open(projiniFileName, 'w')
            #fileobj.write(currentDirectory + '\n')
            #fileobj.write('{}'.format(now) + '\n')
            #fileobj.write(userName + '\n')
            #fileobj.write("{}".format(pnumber) + '\n')
            #fileobj.write("{}".format(nmons) + '\n')
            #for m in range(nmons):
            #    mg = screen.get_monitor_geometry(m)
            #    msize = "%d,%d,%d" % (m, mg.width, mg.height)
            #    monitors.append(msize)
            #fileobj.write("{}".format(monitors) + '\n')
            #fileobj.write("{}".format(lastUsedMonitor))
            #fileobj.write("{}".format(lastX))
            #fileobj.write("{}".format(lastY))
            #fileobj.write('\nPreviously:\n')
            #fileobj.write(location)
            #fileobj.write(lastOpened)
            #fileobj.write(lastUserName)
            #fileobj.write(lastpnumber)
            #fileobj.write(lastQtyMons)
            #fileobj.write(lastMonitorsizes)
            #fileobj.write(lastUsedMonitor)
            #fileobj.write(lastX)
            #fileobj.write(lastY)
            #fileobj.close()

    def ConfigSectionMap(self, section):
        dict1 = {}
        options = Config.options(section)
        for option in options:
            try:
                dict1[option] = Config.get(section, option)
                if dict1[option] == -1:
                    print("skip: %s" % option)
            except:
                print("exception on %s!" % option)
                dict1[option] = None
        return dict1

    def connectPySQL(self):
        connectDb = sqlite3.connect('S32.db')
        connectDb.row_factory = sqlite3.Row
        app.processEvents()
        return connectDb

    def createPySQL(self):
        Qry_Create_tbl_BankStatementSettings = ('''Create Table BankSettings (RecordID int IDENTITY(1,1) PRIMARY KEY, BankName TEXT DEFAULT 0 NOT NULL, StatementSettingName TEXT DEFAULT 0 NOT NULL,StatementSettingLocation TEXT DEFAULT 0 NOT NULL)''')
        queryCursor = self.connectPySQL()
        queryCursor.execute(Qry_Create_tbl_BankStatementSettings)
        queryCursor.commit()
        queryCursor.close()

        return

    def populate_BankStatementSettings(self):
        queryCursor = self.connectPySQL()

        #Nedbank:
        QryNedbank_StatementField_AccNr = ("INSERT INTO BankSettings (BankName, StatementSettingName, StatementSettingLocation) VALUES ('Nedbank', 'AccountNumberField', 'B2')")
        queryCursor.execute(QryNedbank_StatementField_AccNr)
        queryCursor.commit()

        QryNedbank_StatementField_AccDesc = ("INSERT INTO BankSettings (BankName, StatementSettingName, StatementSettingLocation) VALUES ('Nedbank', 'AccountNumberField', 'B2')")
        queryCursor.execute(QryNedbank_StatementField_AccDesc)
        queryCursor.commit()

        QryNedbank_StatementField_StartRow = ("INSERT INTO BankSettings (BankName, StatementSettingName, StatementSettingLocation) VALUES ('Nedbank', 'TransactionStartRow', 5)")
        queryCursor.execute(QryNedbank_StatementField_StartRow)
        queryCursor.commit()

        QryNedbank_StatementField_Date = ("INSERT INTO BankSettings (BankName, StatementSettingName, StatementSettingLocation) VALUES ('Nedbank', 'TransactionDate', 'A')")
        queryCursor.execute(QryNedbank_StatementField_Date)
        queryCursor.commit()

        QryNedbank_StatementField_Description = ("INSERT INTO BankSettings (BankName, StatementSettingName, StatementSettingLocation) VALUES ('Nedbank', 'TransactionDescription', 'B')")
        queryCursor.execute(QryNedbank_StatementField_Description)
        queryCursor.commit()

        QryNedbank_StatementField_TransactionAmount = ("INSERT INTO BankSettings (BankName, StatementSettingName, StatementSettingLocation) VALUES ('Nedbank', 'TransactionAmount', 'C')")
        queryCursor.execute(QryNedbank_StatementField_TransactionAmount)
        queryCursor.commit()

        QryNedbank_StatementField_TransactionAccountSaldo = ("INSERT INTO BankSettings (BankName, StatementSettingName, StatementSettingLocation) VALUES ('Nedbank', 'TransactionAccountSaldo, 'D')")
        queryCursor.execute(QryNedbank_StatementField_TransactionAccountSaldo)
        queryCursor.commit()

        queryCursor.close()
        return

    def checkBankAccTable(self):
        queryCursor = self.connectPySQL()
        queryCursor.execute("""SELECT COUNT(*) FROM information_schema.tables""")

    def createBankAccTable(self):
        queryCursor = self.connectPySQL()
        queryCursor.execute("""SELECT COUNT(*) FROM information_schema.tables""")

    def gen_rows(stream, max_length=None):
        rows = csv.reader(stream)
        if max_length is None:
            rows = list(rows)
            max_length = max(len(row) for row in rows)
        for row in rows:
            yield row + [None] * (max_length - len(row))


    def openCSV(self):


        names = QtGui.QFileDialog.getOpenFileNames(self, 'Open Files... ', QtCore.QDir.currentPath(),
                                                   'CSV (*.csv)')
        for fname in names:
            print ('Opening CSV file \nPlease wait... ')
            with open(fname) as f:
                #df = pd.read_csv(filename=(str(fname).split('/')[-1:][0]),header=None)
                #acc_df = pd.DataFrame.from_records(list(self.gen_rows(f)))
                acc_df = pd.read_csv(f,header=None)
                print ('Data File opened... ')
                Accnr = (acc_df.iloc[1, 1])
                AccName = (acc_df.iloc[2, 1])
                print ("Acc Nr: {}".format(Accnr))
                print ("Acc Name: {}".format(AccName))
                print ("len: {}".format(len(acc_df)))


        #return dataframe

    def getGlobalVariable(self, variable):
        try:
            connectDB = self.connectPySQL()
            with connectDB:
                queryCursor = connectDB.cursor()
                Qry = ("SELECT VarValue, VariableDescription FROM GlobalVariables WHERE Variable = '{}'".format(
                    variable))
                queryCursor.execute(Qry)
                Qry_result = queryCursor.fetchall()
                queryCursor.close()
                result = []
                for x in Qry_result:
                    result.append(x)
                    # print(x)
                return x
        except:
            pass

    def home(self):
        extractAction = QtGui.QAction(QtGui.QIcon('exit.png'), 'Quit', self)
        extractAction.triggered.connect(self.close_application)

        configAction = QtGui.QAction(QtGui.QIcon('configuration-1.png'), 'Config...', self)

        openAction = QtGui.QAction(QtGui.QIcon('folder_green_open.png'), 'Open...', self)
        openAction.triggered.connect(self.openCSV)

        self.toolBar = self.addToolBar("Extraction")
        self.toolBar.addAction(extractAction)
        self.toolBar.addAction(configAction)
        self.toolBar.addAction(openAction)

    def close_application(self):
        choice = QtGui.QMessageBox.question(self, 'Exit!',"Sure?",QtGui.QMessageBox.Yes | QtGui.QMessageBox.No)
        if choice == QtGui.QMessageBox.Yes:
            print("Exit")
            sys.exit()
        else:
            pass

    def get_fname(self):
        # Fullfilename = QtGui.QFileDialog.getOpenFileName(self, 'Open Files...', QtCore.QDir.currentPath(), 'Excel Spreadsheet (*.xlsx)')
        # Fullfilename = QtGui.QFileDialog.getOpenFileName(self, 'Select file').resolve()
        filter = 'CSV (*.csv)' #'Excel Spreadsheet (*.xlsx)'  # filter = "TXT (*.txt);;PDF (*.pdf)"
        file_name = QtGui.QFileDialog()
        file_name.setFileMode(QFileDialog.ExistingFile)
        file_location = self.QtCore.QDir.currentPath()  # file_location = "C\\Desktop"
        file_openMsg = "Open file..."
        Fullfilename = file_name.getOpenFileNamesAndFilter(self, file_openMsg, QtCore.QDir.currentPath(), filter)
        return Fullfilename

    def get_fnames(self):
        filter = 'CSV (*.csv)' # filter = "TXT (*.txt);;PDF (*.pdf)"
        file_name = QtGui.QFileDialog()
        file_name.setFileMode(QFileDialog.ExistingFiles)
        file_openMsg = "Open files..."
        Fullfilenames = file_name.getOpenFileNamesAndFilter(self, file_openMsg, QtCore.QDir.currentPath(), filter)
        return Fullfilenames

        # Open Excel File, and start importing data to SQL Server

    def openWorkbook(self):

        wb = Workbook()

        names = QtGui.QFileDialog.getOpenFileNames(self, 'Open Files... ', QtCore.QDir.currentPath(),
                                                   'CSV (*.csv)')
        for fname in names:
            print ('Opening Excel Data file \nPlease wait... ')
            excel_document = load_workbook(filename=(str(fname).split('/')[-1:][0]))
            print ('Data File opened... ')
            allSheetNames = excel_document.get_sheet_names()
            print ("allSheetNames: {}".format(allSheetNames))
            '''
            for sheetname in allSheetNames:
                wb.active = excel_document[sheetname]
                # activesheet = excel_document[sheetname]
                print (sheetname)
                if 'CV 15 Daily Qualityies' in sheetname:
                    print('Starting with Daily Qualities: {}'.format(sheetname))
                    # ws = wb["CV 15 Daily Qualityies"]
                    self.transfer_OC_quality_Data_to_PySQL(excel_document, sheetname, '150')
                    print('Done')
                elif 'CV 15.5 Daily Qualityies' in sheetname:
                    print('Starting with Daily Qualities: {}'.format(sheetname))
                    # ToDo: include other tables also
                    # self.transfer_OC_quality_Data_to_PySQL(excel_document, sheetname, '155')
                    print('Done')
                elif 'CV 15.8 Daily Qualityies' in sheetname:
                    print('Starting with Daily Qualities: {}'.format(sheetname))
                    # ToDo: include other tables also
                    # self.transfer_OC_quality_Data_to_PySQL(excel_document, sheetname, '158')
                    print('Done')
                elif 'CV 16 Daily Qualityies' in sheetname:
                    print('Starting with Daily Qualities: {}'.format(sheetname))
                    # ToDo: include other tables also
                    # self.transfer_OC_quality_Data_to_PySQL(excel_document, sheetname, '160')
                    print('Done')
                elif 'UG 6 Daily' in sheetname:
                    print('Starting with Daily Qualities: {}'.format(sheetname))
                    self.transfer_UG_Data_to_PySQL(excel_document, sheetname, addedMoisture)
                    print('Done')
                elif "CV 15 Daily" in sheetname:
                    print ('Starting with: {}'.format(sheetname))
                    self.transfer_OC_Data_to_PySQL(excel_document, sheetname, addedMoisture, 'tbl_150')
                    print ('Done')
                elif "CV 15.5 Daily" in sheetname:
                    print ('Starting with: {}'.format(sheetname))
                    # ToDo: include other tables also
                    # self.transfer_OC_Data_to_PySQL(excel_document, sheetname, addedMoisture, 'tbl_155')
                    print('Done')
                elif "CV 15.8 Daily" in sheetname:
                    print ('Starting with: {}'.format(sheetname))
                    # ToDo: include other tables also
                    # self.transfer_OC_Data_to_PySQL(excel_document, sheetname, addedMoisture, 'tbl_158')
                    print('Done')
                elif "CV 16 Daily" in sheetname:
                    print ('Starting with: {}'.format(sheetname))
                    # ToDo: include other tables also
                    # self.transfer_OC_Data_to_PySQL(excel_document, sheetname, addedMoisture, 'tbl_160')
                    print('Done')
                else:
                    pass
                    '''
            return


    def ConnectPySQL(self):
        # connectDb = MySQLdb.connect(host="localhost",  # your host, usually localhost
        #                            user="S32",  # your username
        #                            passwd="1qaz2wsx3edc",  # your password
        #                            db="S32")  # name of the data base
        # connectDb = sqlite3.connect('S32.db')
        # return connectDb

        connectDb = sqlite3.connect('BankStatements.db')
        connectDb.row_factory = sqlite3.Row
        return connectDb


    def CreatePySQL(self):
        print ("Clearing database...")
        tables = (
        'tbl_150', 'tbl_155', 'tbl_158', 'tbl_160', 'S2_150', 'S4_150', 'S2_155', 'S4_155', 'S2_158', 'S4_158', 'S2_160',
        'S4_160')
        # tables = ('tbl_150', 'S2_150', 'S4_150')
        dbcon = self.connectPySQL()
        for tablename in tables:
            self.dropPySQLTableExists(dbcon, tablename)
        self.dropPySQLTableExists(dbcon, 'GlobalVariables')
        dbcon.close()

        queryCursor = self.connectPySQL()
        print ("Done\nCreate Table 150...")
        Q_Create_tbl_150 = ('''Create Table tbl_150 (RecordID INTEGER IDENTITY(1,1) PRIMARY KEY,
                           recordDate TEXT DEFAULT 0 NOT NULL,
                           topSoil REAL DEFAULT 0 NOT NULL,
                           softs REAL DEFAULT 0 NOT NULL,
                           overBurden REAL DEFAULT 0 NOT NULL,
                           midBurden REAL DEFAULT 0 NOT NULL,
                           totalWaste REAL DEFAULT 0 NOT NULL,
                           s2_r_ton REAL DEFAULT 0 NOT NULL,
                           s2_ad_ton REAL DEFAULT 0 NOT NULL,
                           s2_r_cv REAL DEFAULT 0 NOT NULL,
                           s2_ad_cv REAL DEFAULT 0 NOT NULL,
                           s2_r_as REAL DEFAULT 0 NOT NULL,
                           s2_ad_as REAL DEFAULT 0 NOT NULL,
                           s2_r_vm REAL DEFAULT 0 NOT NULL,
                           s2_ad_vm REAL DEFAULT 0 NOT NULL,
                           s2_r_im REAL DEFAULT 0 NOT NULL,
                           s2_ad_im REAL DEFAULT 0 NOT NULL,
                           s2_r_ts REAL DEFAULT 0 NOT NULL,
                           s2_ad_ts REAL DEFAULT 0 NOT NULL,
                           s2_w_ad_ton REAL DEFAULT 0 NOT NULL,
                           s2_w_ad_rd REAL DEFAULT 0 NOT NULL,
                           s2_discard_ton REAL DEFAULT 0 NOT NULL,
                           s2_discard_cv REAL DEFAULT 0 NOT NULL,
                           s4_r_ton REAL DEFAULT 0 NOT NULL,
                           s4_ad_ton REAL DEFAULT 0 NOT NULL,
                           s4_r_cv REAL DEFAULT 0 NOT NULL,
                           s4_ad_cv REAL DEFAULT 0 NOT NULL,
                           s4_r_as REAL DEFAULT 0 NOT NULL,
                           s4_ad_as REAL DEFAULT 0 NOT NULL,
                           s4_r_vm REAL DEFAULT 0 NOT NULL,
                           s4_ad_vm REAL DEFAULT 0 NOT NULL,
                           s4_r_im REAL DEFAULT 0 NOT NULL,
                           s4_ad_im REAL DEFAULT 0 NOT NULL,
                           s4_r_ts REAL DEFAULT 0 NOT NULL,
                           s4_ad_ts REAL DEFAULT 0 NOT NULL,
                           s4_w_ad_ton REAL DEFAULT 0 NOT NULL,
                           s4_w_ad_rd REAL DEFAULT 0 NOT NULL,
                           s4_discard_ton REAL DEFAULT 0 NOT NULL,
                           s4_discard_cv REAL DEFAULT 0 NOT NULL,
                           ug6_r_ton REAL DEFAULT 0 NOT NULL,
                           ug6_ad_ton REAL DEFAULT 0 NOT NULL,
                           ug6_r_cv REAL DEFAULT 0 NOT NULL,
                           ug6_ad_cv REAL DEFAULT 0 NOT NULL,
                           ug6_r_as REAL DEFAULT 0 NOT NULL,
                           ug6_ad_as REAL DEFAULT 0 NOT NULL,
                           ug6_r_vm REAL DEFAULT 0 NOT NULL,
                           ug6_ad_vm REAL DEFAULT 0 NOT NULL,
                           ug6_r_im REAL DEFAULT 0 NOT NULL,
                           ug6_ad_im REAL DEFAULT 0 NOT NULL,
                           ug6_r_ts REAL DEFAULT 0 NOT NULL,
                           ug6_ad_ts REAL DEFAULT 0 NOT NULL,
                           rom_ad_cv REAL DEFAULT 0 NOT NULL,
                           rom_ad_as REAL DEFAULT 0 NOT NULL,
                           rom_ad_vm REAL DEFAULT 0 NOT NULL,
                           rom_ad_ton REAL DEFAULT 0 NOT NULL,
                           rom_w_ad_ton REAL DEFAULT 0 NOT NULL,
                           rom_w_ad_cv REAL DEFAULT 0 NOT NULL,
                           rom_w_ad_as REAL DEFAULT 0 NOT NULL,
                           rom_w_ad_vm REAL DEFAULT 0 NOT NULL,
                           discard_ad_ton REAL DEFAULT 0 NOT NULL,
                           discard_ad_cv REAL DEFAULT 0 NOT NULL,
                           discard_ad_energy REAL DEFAULT 0 NOT NULL,
                            total_ad_ton REAL DEFAULT 0 NOT NULL,
                            total_ad_cv REAL DEFAULT 0 NOT NULL,
                            total_ad_as REAL DEFAULT 0 NOT NULL
                           )''')
        queryCursor.execute(Q_Create_tbl_150)
        print("Done\nCreate Table 155...")
        Q_Create_tbl_155 = ('''Create Table tbl_155 (RecordID INTEGER IDENTITY(1,1) PRIMARY KEY,
                           recordDate TEXT DEFAULT 0 NOT NULL,
                           topSoil REAL DEFAULT 0 NOT NULL,
                           softs REAL DEFAULT 0 NOT NULL,
                           overBurden REAL DEFAULT 0 NOT NULL,
                           midBurden REAL DEFAULT 0 NOT NULL,
                           totalWaste REAL DEFAULT 0 NOT NULL,
                           s2_r_ton REAL DEFAULT 0 NOT NULL,
                           s2_ad_ton REAL DEFAULT 0 NOT NULL,
                           s2_r_cv REAL DEFAULT 0 NOT NULL,
                           s2_ad_cv REAL DEFAULT 0 NOT NULL,
                           s2_r_as REAL DEFAULT 0 NOT NULL,
                           s2_ad_as REAL DEFAULT 0 NOT NULL,
                           s2_r_vm REAL DEFAULT 0 NOT NULL,
                           s2_ad_vm REAL DEFAULT 0 NOT NULL,
                           s2_r_im REAL DEFAULT 0 NOT NULL,
                           s2_ad_im REAL DEFAULT 0 NOT NULL,
                           s2_r_ts REAL DEFAULT 0 NOT NULL,
                           s2_ad_ts REAL DEFAULT 0 NOT NULL,
                           s2_w_ad_ton REAL DEFAULT 0 NOT NULL,
                           s2_w_ad_rd REAL DEFAULT 0 NOT NULL,
                           s2_discard_ton REAL DEFAULT 0 NOT NULL,
                           s2_discard_cv REAL DEFAULT 0 NOT NULL,
                           s4_r_ton REAL DEFAULT 0 NOT NULL,
                           s4_ad_ton REAL DEFAULT 0 NOT NULL,
                           s4_r_cv REAL DEFAULT 0 NOT NULL,
                           s4_ad_cv REAL DEFAULT 0 NOT NULL,
                           s4_r_as REAL DEFAULT 0 NOT NULL,
                           s4_ad_as REAL DEFAULT 0 NOT NULL,
                           s4_r_vm REAL DEFAULT 0 NOT NULL,
                           s4_ad_vm REAL DEFAULT 0 NOT NULL,
                           s4_r_im REAL DEFAULT 0 NOT NULL,
                           s4_ad_im REAL DEFAULT 0 NOT NULL,
                           s4_r_ts REAL DEFAULT 0 NOT NULL,
                           s4_ad_ts REAL DEFAULT 0 NOT NULL,
                           s4_w_ad_ton REAL DEFAULT 0 NOT NULL,
                           s4_w_ad_rd REAL DEFAULT 0 NOT NULL,
                           s4_discard_ton REAL DEFAULT 0 NOT NULL,
                           s4_discard_cv REAL DEFAULT 0 NOT NULL,
                           ug6_r_ton REAL DEFAULT 0 NOT NULL,
                           ug6_ad_ton REAL DEFAULT 0 NOT NULL,
                           ug6_r_cv REAL DEFAULT 0 NOT NULL,
                           ug6_ad_cv REAL DEFAULT 0 NOT NULL,
                           ug6_r_as REAL DEFAULT 0 NOT NULL,
                           ug6_ad_as REAL DEFAULT 0 NOT NULL,
                           ug6_r_vm REAL DEFAULT 0 NOT NULL,
                           ug6_ad_vm REAL DEFAULT 0 NOT NULL,
                           ug6_r_im REAL DEFAULT 0 NOT NULL,
                           ug6_ad_im REAL DEFAULT 0 NOT NULL,
                           ug6_r_ts REAL DEFAULT 0 NOT NULL,
                           ug6_ad_ts REAL DEFAULT 0 NOT NULL,
                           rom_ad_cv REAL DEFAULT 0 NOT NULL,
                           rom_ad_as REAL DEFAULT 0 NOT NULL,
                           rom_ad_vm REAL DEFAULT 0 NOT NULL,
                           rom_ad_ton REAL DEFAULT 0 NOT NULL,
                           rom_w_ad_ton REAL DEFAULT 0 NOT NULL,
                           rom_w_ad_cv REAL DEFAULT 0 NOT NULL,
                           rom_w_ad_as REAL DEFAULT 0 NOT NULL,
                           rom_w_ad_vm REAL DEFAULT 0 NOT NULL,
                           discard_ad_ton REAL DEFAULT 0 NOT NULL,
                           discard_ad_cv REAL DEFAULT 0 NOT NULL,
                           discard_ad_energy REAL DEFAULT 0 NOT NULL,
                            total_ad_ton REAL DEFAULT 0 NOT NULL,
                            total_ad_cv REAL DEFAULT 0 NOT NULL,
                            total_ad_as REAL DEFAULT 0 NOT NULL)''')
        queryCursor.execute(Q_Create_tbl_155)
        print("Done\nCreate Table 158...")
        Q_Create_tbl_158 = ('''Create Table tbl_158 (RecordID INTEGER IDENTITY(1,1) PRIMARY KEY,
                           recordDate TEXT DEFAULT 0 NOT NULL,
                           topSoil REAL DEFAULT 0 NOT NULL,
                           softs REAL DEFAULT 0 NOT NULL,
                           overBurden REAL DEFAULT 0 NOT NULL,
                           midBurden REAL DEFAULT 0 NOT NULL,
                           totalWaste REAL DEFAULT 0 NOT NULL,
                           s2_r_ton REAL DEFAULT 0 NOT NULL,
                           s2_ad_ton REAL DEFAULT 0 NOT NULL,
                           s2_r_cv REAL DEFAULT 0 NOT NULL,
                           s2_ad_cv REAL DEFAULT 0 NOT NULL,
                           s2_r_as REAL DEFAULT 0 NOT NULL,
                           s2_ad_as REAL DEFAULT 0 NOT NULL,
                           s2_r_vm REAL DEFAULT 0 NOT NULL,
                           s2_ad_vm REAL DEFAULT 0 NOT NULL,
                           s2_r_im REAL DEFAULT 0 NOT NULL,
                           s2_ad_im REAL DEFAULT 0 NOT NULL,
                           s2_r_ts REAL DEFAULT 0 NOT NULL,
                           s2_ad_ts REAL DEFAULT 0 NOT NULL,
                           s2_w_ad_ton REAL DEFAULT 0 NOT NULL,
                           s2_w_ad_rd REAL DEFAULT 0 NOT NULL,
                           s2_discard_ton REAL DEFAULT 0 NOT NULL,
                           s2_discard_cv REAL DEFAULT 0 NOT NULL,
                           s4_r_ton REAL DEFAULT 0 NOT NULL,
                           s4_ad_ton REAL DEFAULT 0 NOT NULL,
                           s4_r_cv REAL DEFAULT 0 NOT NULL,
                           s4_ad_cv REAL DEFAULT 0 NOT NULL,
                           s4_r_as REAL DEFAULT 0 NOT NULL,
                           s4_ad_as REAL DEFAULT 0 NOT NULL,
                           s4_r_vm REAL DEFAULT 0 NOT NULL,
                           s4_ad_vm REAL DEFAULT 0 NOT NULL,
                           s4_r_im REAL DEFAULT 0 NOT NULL,
                           s4_ad_im REAL DEFAULT 0 NOT NULL,
                           s4_r_ts REAL DEFAULT 0 NOT NULL,
                           s4_ad_ts REAL DEFAULT 0 NOT NULL,
                           s4_w_ad_ton REAL DEFAULT 0 NOT NULL,
                           s4_w_ad_rd REAL DEFAULT 0 NOT NULL,
                           s4_discard_ton REAL DEFAULT 0 NOT NULL,
                           s4_discard_cv REAL DEFAULT 0 NOT NULL,
                           ug6_r_ton REAL DEFAULT 0 NOT NULL,
                           ug6_ad_ton REAL DEFAULT 0 NOT NULL,
                           ug6_r_cv REAL DEFAULT 0 NOT NULL,
                           ug6_ad_cv REAL DEFAULT 0 NOT NULL,
                           ug6_r_as REAL DEFAULT 0 NOT NULL,
                           ug6_ad_as REAL DEFAULT 0 NOT NULL,
                           ug6_r_vm REAL DEFAULT 0 NOT NULL,
                           ug6_ad_vm REAL DEFAULT 0 NOT NULL,
                           ug6_r_im REAL DEFAULT 0 NOT NULL,
                           ug6_ad_im REAL DEFAULT 0 NOT NULL,
                           ug6_r_ts REAL DEFAULT 0 NOT NULL,
                           ug6_ad_ts REAL DEFAULT 0 NOT NULL,
                           rom_ad_cv REAL DEFAULT 0 NOT NULL,
                           rom_ad_as REAL DEFAULT 0 NOT NULL,
                           rom_ad_vm REAL DEFAULT 0 NOT NULL,
                           rom_ad_ton REAL DEFAULT 0 NOT NULL,
                           rom_w_ad_ton REAL DEFAULT 0 NOT NULL,
                           rom_w_ad_cv REAL DEFAULT 0 NOT NULL,
                           rom_w_ad_as REAL DEFAULT 0 NOT NULL,
                           rom_w_ad_vm REAL DEFAULT 0 NOT NULL,
                           discard_ad_ton REAL DEFAULT 0 NOT NULL,
                           discard_ad_cv REAL DEFAULT 0 NOT NULL,
                           discard_ad_energy REAL DEFAULT 0 NOT NULL,
                            total_ad_ton REAL DEFAULT 0 NOT NULL,
                            total_ad_cv REAL DEFAULT 0 NOT NULL,
                            total_ad_as REAL DEFAULT 0 NOT NULL)''')
        queryCursor.execute(Q_Create_tbl_158)
        print("Done\nCreate Table 160...")
        Q_Create_tbl_160 = ('''Create Table tbl_160 (RecordID INTEGER IDENTITY(1,1) PRIMARY KEY,
                           recordDate TEXT DEFAULT 0 NOT NULL,
                           topSoil REAL DEFAULT 0 NOT NULL,
                           softs REAL DEFAULT 0 NOT NULL,
                           overBurden REAL DEFAULT 0 NOT NULL,
                           midBurden REAL DEFAULT 0 NOT NULL,
                           totalWaste REAL DEFAULT 0 NOT NULL,
                           s2_r_ton REAL DEFAULT 0 NOT NULL,
                           s2_ad_ton REAL DEFAULT 0 NOT NULL,
                           s2_r_cv REAL DEFAULT 0 NOT NULL,
                           s2_ad_cv REAL DEFAULT 0 NOT NULL,
                           s2_r_as REAL DEFAULT 0 NOT NULL,
                           s2_ad_as REAL DEFAULT 0 NOT NULL,
                           s2_r_vm REAL DEFAULT 0 NOT NULL,
                           s2_ad_vm REAL DEFAULT 0 NOT NULL,
                           s2_r_im REAL DEFAULT 0 NOT NULL,
                           s2_ad_im REAL DEFAULT 0 NOT NULL,
                           s2_r_ts REAL DEFAULT 0 NOT NULL,
                           s2_ad_ts REAL DEFAULT 0 NOT NULL,
                           s2_w_ad_ton REAL DEFAULT 0 NOT NULL,
                           s2_w_ad_rd REAL DEFAULT 0 NOT NULL,
                           s2_discard_ton REAL DEFAULT 0 NOT NULL,
                           s2_discard_cv REAL DEFAULT 0 NOT NULL,
                           s4_r_ton REAL DEFAULT 0 NOT NULL,
                           s4_ad_ton REAL DEFAULT 0 NOT NULL,
                           s4_r_cv REAL DEFAULT 0 NOT NULL,
                           s4_ad_cv REAL DEFAULT 0 NOT NULL,
                           s4_r_as REAL DEFAULT 0 NOT NULL,
                           s4_ad_as REAL DEFAULT 0 NOT NULL,
                           s4_r_vm REAL DEFAULT 0 NOT NULL,
                           s4_ad_vm REAL DEFAULT 0 NOT NULL,
                           s4_r_im REAL DEFAULT 0 NOT NULL,
                           s4_ad_im REAL DEFAULT 0 NOT NULL,
                           s4_r_ts REAL DEFAULT 0 NOT NULL,
                           s4_ad_ts REAL DEFAULT 0 NOT NULL,
                           s4_w_ad_ton REAL DEFAULT 0 NOT NULL,
                           s4_w_ad_rd REAL DEFAULT 0 NOT NULL,
                           s4_discard_ton REAL DEFAULT 0 NOT NULL,
                           s4_discard_cv REAL DEFAULT 0 NOT NULL,
                           ug6_r_ton REAL DEFAULT 0 NOT NULL,
                           ug6_ad_ton REAL DEFAULT 0 NOT NULL,
                           ug6_r_cv REAL DEFAULT 0 NOT NULL,
                           ug6_ad_cv REAL DEFAULT 0 NOT NULL,
                           ug6_r_as REAL DEFAULT 0 NOT NULL,
                           ug6_ad_as REAL DEFAULT 0 NOT NULL,
                           ug6_r_vm REAL DEFAULT 0 NOT NULL,
                           ug6_ad_vm REAL DEFAULT 0 NOT NULL,
                           ug6_r_im REAL DEFAULT 0 NOT NULL,
                           ug6_ad_im REAL DEFAULT 0 NOT NULL,
                           ug6_r_ts REAL DEFAULT 0 NOT NULL,
                           ug6_ad_ts REAL DEFAULT 0 NOT NULL,
                           rom_ad_cv REAL DEFAULT 0 NOT NULL,
                           rom_ad_as REAL DEFAULT 0 NOT NULL,
                           rom_ad_vm REAL DEFAULT 0 NOT NULL,
                           rom_ad_ton REAL DEFAULT 0 NOT NULL,
                           rom_w_ad_ton REAL DEFAULT 0 NOT NULL,
                           rom_w_ad_cv REAL DEFAULT 0 NOT NULL,
                           rom_w_ad_as REAL DEFAULT 0 NOT NULL,
                           rom_w_ad_vm REAL DEFAULT 0 NOT NULL,
                           discard_ad_ton REAL DEFAULT 0 NOT NULL,
                           discard_ad_cv REAL DEFAULT 0 NOT NULL,
                           discard_ad_energy REAL DEFAULT 0 NOT NULL,
                            total_ad_ton REAL DEFAULT 0 NOT NULL,
                            total_ad_cv REAL DEFAULT 0 NOT NULL,
                            total_ad_as REAL DEFAULT 0 NOT NULL)''')
        queryCursor.execute(Q_Create_tbl_160)
        print("Done\nCreate Table S2 150...")
        Q_Create_S2_150 = ('''Create Table S2_150 (RecordID int IDENTITY(1,1) PRIMARY KEY,
                            recordDate TEXT DEFAULT 0 NOT NULL,
                            DA REAL DEFAULT 0 NOT NULL, DC REAL DEFAULT 0 NOT NULL,
                            DM REAL DEFAULT 0 NOT NULL, DS REAL DEFAULT 0 NOT NULL,
                            DV REAL DEFAULT 0 NOT NULL, DY REAL DEFAULT 0 NOT NULL,
                            EA REAL DEFAULT 0 NOT NULL, EC REAL DEFAULT 0 NOT NULL,
                            EM REAL DEFAULT 0 NOT NULL, ES REAL DEFAULT 0 NOT NULL,
                            EV REAL DEFAULT 0 NOT NULL, EY REAL DEFAULT 0 NOT NULL,
                            FA REAL DEFAULT 0 NOT NULL, FC REAL DEFAULT 0 NOT NULL,
                            FM REAL DEFAULT 0 NOT NULL, FS REAL DEFAULT 0 NOT NULL,
                            FV REAL DEFAULT 0 NOT NULL, FY REAL DEFAULT 0 NOT NULL,
                            GA REAL DEFAULT 0 NOT NULL, GC REAL DEFAULT 0 NOT NULL,
                            GM REAL DEFAULT 0 NOT NULL, GS REAL DEFAULT 0 NOT NULL,
                            GV REAL DEFAULT 0 NOT NULL, GY REAL DEFAULT 0 NOT NULL,
                            HA REAL DEFAULT 0 NOT NULL, HC REAL DEFAULT 0 NOT NULL,
                            HM REAL DEFAULT 0 NOT NULL, HS REAL DEFAULT 0 NOT NULL,
                            HV REAL DEFAULT 0 NOT NULL, HY REAL DEFAULT 0 NOT NULL,
                            JA REAL DEFAULT 0 NOT NULL, JC REAL DEFAULT 0 NOT NULL,
                            JM REAL DEFAULT 0 NOT NULL, JS REAL DEFAULT 0 NOT NULL,
                            JV REAL DEFAULT 0 NOT NULL, JY REAL DEFAULT 0 NOT NULL,
                            KA REAL DEFAULT 0 NOT NULL, KC REAL DEFAULT 0 NOT NULL,
                            KM REAL DEFAULT 0 NOT NULL, KS REAL DEFAULT 0 NOT NULL,
                            KV REAL DEFAULT 0 NOT NULL, KY REAL DEFAULT 0 NOT NULL,
                            LA REAL DEFAULT 0 NOT NULL, LC REAL DEFAULT 0 NOT NULL,
                            LM REAL DEFAULT 0 NOT NULL, LS REAL DEFAULT 0 NOT NULL,
                            LV REAL DEFAULT 0 NOT NULL, LY REAL DEFAULT 0 NOT NULL,
                            MA REAL DEFAULT 0 NOT NULL, MC REAL DEFAULT 0 NOT NULL,
                            MM REAL DEFAULT 0 NOT NULL, MS REAL DEFAULT 0 NOT NULL,
                            MV REAL DEFAULT 0 NOT NULL, MY REAL DEFAULT 0 NOT NULL,
                            NA REAL DEFAULT 0 NOT NULL, NC REAL DEFAULT 0 NOT NULL,
                            NM REAL DEFAULT 0 NOT NULL, NS REAL DEFAULT 0 NOT NULL,
                            NV REAL DEFAULT 0 NOT NULL, NY REAL DEFAULT 0 NOT NULL)''')
        queryCursor.execute(Q_Create_S2_150)
        print("Done\nCreate Table S2 155...")
        Q_Create_S2_155 = ('''Create Table S2_155 (RecordID int IDENTITY(1,1) PRIMARY KEY,
                            recordDate TEXT DEFAULT 0 NOT NULL,
                            DA REAL DEFAULT 0 NOT NULL, DC REAL DEFAULT 0 NOT NULL,
                            DM REAL DEFAULT 0 NOT NULL, DS REAL DEFAULT 0 NOT NULL,
                            DV REAL DEFAULT 0 NOT NULL, DY REAL DEFAULT 0 NOT NULL,
                            EA REAL DEFAULT 0 NOT NULL, EC REAL DEFAULT 0 NOT NULL,
                            EM REAL DEFAULT 0 NOT NULL, ES REAL DEFAULT 0 NOT NULL,
                            EV REAL DEFAULT 0 NOT NULL, EY REAL DEFAULT 0 NOT NULL,
                            FA REAL DEFAULT 0 NOT NULL, FC REAL DEFAULT 0 NOT NULL,
                            FM REAL DEFAULT 0 NOT NULL, FS REAL DEFAULT 0 NOT NULL,
                            FV REAL DEFAULT 0 NOT NULL, FY REAL DEFAULT 0 NOT NULL,
                            GA REAL DEFAULT 0 NOT NULL, GC REAL DEFAULT 0 NOT NULL,
                            GM REAL DEFAULT 0 NOT NULL, GS REAL DEFAULT 0 NOT NULL,
                            GV REAL DEFAULT 0 NOT NULL, GY REAL DEFAULT 0 NOT NULL,
                            HA REAL DEFAULT 0 NOT NULL, HC REAL DEFAULT 0 NOT NULL,
                            HM REAL DEFAULT 0 NOT NULL, HS REAL DEFAULT 0 NOT NULL,
                            HV REAL DEFAULT 0 NOT NULL, HY REAL DEFAULT 0 NOT NULL,
                            JA REAL DEFAULT 0 NOT NULL, JC REAL DEFAULT 0 NOT NULL,
                            JM REAL DEFAULT 0 NOT NULL, JS REAL DEFAULT 0 NOT NULL,
                            JV REAL DEFAULT 0 NOT NULL, JY REAL DEFAULT 0 NOT NULL,
                            KA REAL DEFAULT 0 NOT NULL, KC REAL DEFAULT 0 NOT NULL,
                            KM REAL DEFAULT 0 NOT NULL, KS REAL DEFAULT 0 NOT NULL,
                            KV REAL DEFAULT 0 NOT NULL, KY REAL DEFAULT 0 NOT NULL,
                            LA REAL DEFAULT 0 NOT NULL, LC REAL DEFAULT 0 NOT NULL,
                            LM REAL DEFAULT 0 NOT NULL, LS REAL DEFAULT 0 NOT NULL,
                            LV REAL DEFAULT 0 NOT NULL, LY REAL DEFAULT 0 NOT NULL,
                            MA REAL DEFAULT 0 NOT NULL, MC REAL DEFAULT 0 NOT NULL,
                            MM REAL DEFAULT 0 NOT NULL, MS REAL DEFAULT 0 NOT NULL,
                            MV REAL DEFAULT 0 NOT NULL, MY REAL DEFAULT 0 NOT NULL,
                            NA REAL DEFAULT 0 NOT NULL, NC REAL DEFAULT 0 NOT NULL,
                            NM REAL DEFAULT 0 NOT NULL, NS REAL DEFAULT 0 NOT NULL,
                            NV REAL DEFAULT 0 NOT NULL, NY REAL DEFAULT 0 NOT NULL)''')
        queryCursor.execute(Q_Create_S2_155)
        print("Done\nCreate Table S2 158...")
        Q_Create_S2_158 = ('''Create Table S2_158 (RecordID int IDENTITY(1,1) PRIMARY KEY,
                            recordDate TEXT DEFAULT 0 NOT NULL,
                            DA REAL DEFAULT 0 NOT NULL, DC REAL DEFAULT 0 NOT NULL,
                            DM REAL DEFAULT 0 NOT NULL, DS REAL DEFAULT 0 NOT NULL,
                            DV REAL DEFAULT 0 NOT NULL, DY REAL DEFAULT 0 NOT NULL,
                            EA REAL DEFAULT 0 NOT NULL, EC REAL DEFAULT 0 NOT NULL,
                            EM REAL DEFAULT 0 NOT NULL, ES REAL DEFAULT 0 NOT NULL,
                            EV REAL DEFAULT 0 NOT NULL, EY REAL DEFAULT 0 NOT NULL,
                            FA REAL DEFAULT 0 NOT NULL, FC REAL DEFAULT 0 NOT NULL,
                            FM REAL DEFAULT 0 NOT NULL, FS REAL DEFAULT 0 NOT NULL,
                            FV REAL DEFAULT 0 NOT NULL, FY REAL DEFAULT 0 NOT NULL,
                            GA REAL DEFAULT 0 NOT NULL, GC REAL DEFAULT 0 NOT NULL,
                            GM REAL DEFAULT 0 NOT NULL, GS REAL DEFAULT 0 NOT NULL,
                            GV REAL DEFAULT 0 NOT NULL, GY REAL DEFAULT 0 NOT NULL,
                            HA REAL DEFAULT 0 NOT NULL, HC REAL DEFAULT 0 NOT NULL,
                            HM REAL DEFAULT 0 NOT NULL, HS REAL DEFAULT 0 NOT NULL,
                            HV REAL DEFAULT 0 NOT NULL, HY REAL DEFAULT 0 NOT NULL,
                            JA REAL DEFAULT 0 NOT NULL, JC REAL DEFAULT 0 NOT NULL,
                            JM REAL DEFAULT 0 NOT NULL, JS REAL DEFAULT 0 NOT NULL,
                            JV REAL DEFAULT 0 NOT NULL, JY REAL DEFAULT 0 NOT NULL,
                            KA REAL DEFAULT 0 NOT NULL, KC REAL DEFAULT 0 NOT NULL,
                            KM REAL DEFAULT 0 NOT NULL, KS REAL DEFAULT 0 NOT NULL,
                            KV REAL DEFAULT 0 NOT NULL, KY REAL DEFAULT 0 NOT NULL,
                            LA REAL DEFAULT 0 NOT NULL, LC REAL DEFAULT 0 NOT NULL,
                            LM REAL DEFAULT 0 NOT NULL, LS REAL DEFAULT 0 NOT NULL,
                            LV REAL DEFAULT 0 NOT NULL, LY REAL DEFAULT 0 NOT NULL,
                            MA REAL DEFAULT 0 NOT NULL, MC REAL DEFAULT 0 NOT NULL,
                            MM REAL DEFAULT 0 NOT NULL, MS REAL DEFAULT 0 NOT NULL,
                            MV REAL DEFAULT 0 NOT NULL, MY REAL DEFAULT 0 NOT NULL,
                            NA REAL DEFAULT 0 NOT NULL, NC REAL DEFAULT 0 NOT NULL,
                            NM REAL DEFAULT 0 NOT NULL, NS REAL DEFAULT 0 NOT NULL,
                            NV REAL DEFAULT 0 NOT NULL, NY REAL DEFAULT 0 NOT NULL)''')
        queryCursor.execute(Q_Create_S2_158)
        print("Done\nCreate Table S2 160...")
        Q_Create_S2_160 = ('''Create Table S2_160 (RecordID int IDENTITY(1,1) PRIMARY KEY,
                            recordDate TEXT DEFAULT 0 NOT NULL,
                            DA REAL DEFAULT 0 NOT NULL, DC REAL DEFAULT 0 NOT NULL,
                            DM REAL DEFAULT 0 NOT NULL, DS REAL DEFAULT 0 NOT NULL,
                            DV REAL DEFAULT 0 NOT NULL, DY REAL DEFAULT 0 NOT NULL,
                            EA REAL DEFAULT 0 NOT NULL, EC REAL DEFAULT 0 NOT NULL,
                            EM REAL DEFAULT 0 NOT NULL, ES REAL DEFAULT 0 NOT NULL,
                            EV REAL DEFAULT 0 NOT NULL, EY REAL DEFAULT 0 NOT NULL,
                            FA REAL DEFAULT 0 NOT NULL, FC REAL DEFAULT 0 NOT NULL,
                            FM REAL DEFAULT 0 NOT NULL, FS REAL DEFAULT 0 NOT NULL,
                            FV REAL DEFAULT 0 NOT NULL, FY REAL DEFAULT 0 NOT NULL,
                            GA REAL DEFAULT 0 NOT NULL, GC REAL DEFAULT 0 NOT NULL,
                            GM REAL DEFAULT 0 NOT NULL, GS REAL DEFAULT 0 NOT NULL,
                            GV REAL DEFAULT 0 NOT NULL, GY REAL DEFAULT 0 NOT NULL,
                            HA REAL DEFAULT 0 NOT NULL, HC REAL DEFAULT 0 NOT NULL,
                            HM REAL DEFAULT 0 NOT NULL, HS REAL DEFAULT 0 NOT NULL,
                            HV REAL DEFAULT 0 NOT NULL, HY REAL DEFAULT 0 NOT NULL,
                            JA REAL DEFAULT 0 NOT NULL, JC REAL DEFAULT 0 NOT NULL,
                            JM REAL DEFAULT 0 NOT NULL, JS REAL DEFAULT 0 NOT NULL,
                            JV REAL DEFAULT 0 NOT NULL, JY REAL DEFAULT 0 NOT NULL,
                            KA REAL DEFAULT 0 NOT NULL, KC REAL DEFAULT 0 NOT NULL,
                            KM REAL DEFAULT 0 NOT NULL, KS REAL DEFAULT 0 NOT NULL,
                            KV REAL DEFAULT 0 NOT NULL, KY REAL DEFAULT 0 NOT NULL,
                            LA REAL DEFAULT 0 NOT NULL, LC REAL DEFAULT 0 NOT NULL,
                            LM REAL DEFAULT 0 NOT NULL, LS REAL DEFAULT 0 NOT NULL,
                            LV REAL DEFAULT 0 NOT NULL, LY REAL DEFAULT 0 NOT NULL,
                            MA REAL DEFAULT 0 NOT NULL, MC REAL DEFAULT 0 NOT NULL,
                            MM REAL DEFAULT 0 NOT NULL, MS REAL DEFAULT 0 NOT NULL,
                            MV REAL DEFAULT 0 NOT NULL, MY REAL DEFAULT 0 NOT NULL,
                            NA REAL DEFAULT 0 NOT NULL, NC REAL DEFAULT 0 NOT NULL,
                            NM REAL DEFAULT 0 NOT NULL, NS REAL DEFAULT 0 NOT NULL,
                            NV REAL DEFAULT 0 NOT NULL, NY REAL DEFAULT 0 NOT NULL)''')
        queryCursor.execute(Q_Create_S2_160)
        print("Done\nCreate Table S4 150...")
        Q_Create_S4_150 = ('''Create Table S4_150 (RecordID int IDENTITY(1,1) PRIMARY KEY,
                            recordDate TEXT DEFAULT 0 NOT NULL,
                            DA REAL DEFAULT 0 NOT NULL, DC REAL DEFAULT 0 NOT NULL,
                            DM REAL DEFAULT 0 NOT NULL, DS REAL DEFAULT 0 NOT NULL,
                            DV REAL DEFAULT 0 NOT NULL, DY REAL DEFAULT 0 NOT NULL,
                            EA REAL DEFAULT 0 NOT NULL, EC REAL DEFAULT 0 NOT NULL,
                            EM REAL DEFAULT 0 NOT NULL, ES REAL DEFAULT 0 NOT NULL,
                            EV REAL DEFAULT 0 NOT NULL, EY REAL DEFAULT 0 NOT NULL,
                            FA REAL DEFAULT 0 NOT NULL, FC REAL DEFAULT 0 NOT NULL,
                            FM REAL DEFAULT 0 NOT NULL, FS REAL DEFAULT 0 NOT NULL,
                            FV REAL DEFAULT 0 NOT NULL, FY REAL DEFAULT 0 NOT NULL,
                            GA REAL DEFAULT 0 NOT NULL, GC REAL DEFAULT 0 NOT NULL,
                            GM REAL DEFAULT 0 NOT NULL, GS REAL DEFAULT 0 NOT NULL,
                            GV REAL DEFAULT 0 NOT NULL, GY REAL DEFAULT 0 NOT NULL,
                            HA REAL DEFAULT 0 NOT NULL, HC REAL DEFAULT 0 NOT NULL,
                            HM REAL DEFAULT 0 NOT NULL, HS REAL DEFAULT 0 NOT NULL,
                            HV REAL DEFAULT 0 NOT NULL, HY REAL DEFAULT 0 NOT NULL,
                            JA REAL DEFAULT 0 NOT NULL, JC REAL DEFAULT 0 NOT NULL,
                            JM REAL DEFAULT 0 NOT NULL, JS REAL DEFAULT 0 NOT NULL,
                            JV REAL DEFAULT 0 NOT NULL, JY REAL DEFAULT 0 NOT NULL,
                            KA REAL DEFAULT 0 NOT NULL, KC REAL DEFAULT 0 NOT NULL,
                            KM REAL DEFAULT 0 NOT NULL, KS REAL DEFAULT 0 NOT NULL,
                            KV REAL DEFAULT 0 NOT NULL, KY REAL DEFAULT 0 NOT NULL,
                            LA REAL DEFAULT 0 NOT NULL, LC REAL DEFAULT 0 NOT NULL,
                            LM REAL DEFAULT 0 NOT NULL, LS REAL DEFAULT 0 NOT NULL,
                            LV REAL DEFAULT 0 NOT NULL, LY REAL DEFAULT 0 NOT NULL,
                            MA REAL DEFAULT 0 NOT NULL, MC REAL DEFAULT 0 NOT NULL,
                            MM REAL DEFAULT 0 NOT NULL, MS REAL DEFAULT 0 NOT NULL,
                            MV REAL DEFAULT 0 NOT NULL, MY REAL DEFAULT 0 NOT NULL,
                            NA REAL DEFAULT 0 NOT NULL, NC REAL DEFAULT 0 NOT NULL,
                            NM REAL DEFAULT 0 NOT NULL, NS REAL DEFAULT 0 NOT NULL,
                            NV REAL DEFAULT 0 NOT NULL, NY REAL DEFAULT 0 NOT NULL)''')
        queryCursor.execute(Q_Create_S4_150)
        print("Done\nCreate Table S4 155...")
        Q_Create_S4_155 = ('''Create Table S4_155 (RecordID int IDENTITY(1,1) PRIMARY KEY,
                            recordDate TEXT DEFAULT 0 NOT NULL,
                            DA REAL DEFAULT 0 NOT NULL, DC REAL DEFAULT 0 NOT NULL,
                            DM REAL DEFAULT 0 NOT NULL, DS REAL DEFAULT 0 NOT NULL,
                            DV REAL DEFAULT 0 NOT NULL, DY REAL DEFAULT 0 NOT NULL,
                            EA REAL DEFAULT 0 NOT NULL, EC REAL DEFAULT 0 NOT NULL,
                            EM REAL DEFAULT 0 NOT NULL, ES REAL DEFAULT 0 NOT NULL,
                            EV REAL DEFAULT 0 NOT NULL, EY REAL DEFAULT 0 NOT NULL,
                            FA REAL DEFAULT 0 NOT NULL, FC REAL DEFAULT 0 NOT NULL,
                            FM REAL DEFAULT 0 NOT NULL, FS REAL DEFAULT 0 NOT NULL,
                            FV REAL DEFAULT 0 NOT NULL, FY REAL DEFAULT 0 NOT NULL,
                            GA REAL DEFAULT 0 NOT NULL, GC REAL DEFAULT 0 NOT NULL,
                            GM REAL DEFAULT 0 NOT NULL, GS REAL DEFAULT 0 NOT NULL,
                            GV REAL DEFAULT 0 NOT NULL, GY REAL DEFAULT 0 NOT NULL,
                            HA REAL DEFAULT 0 NOT NULL, HC REAL DEFAULT 0 NOT NULL,
                            HM REAL DEFAULT 0 NOT NULL, HS REAL DEFAULT 0 NOT NULL,
                            HV REAL DEFAULT 0 NOT NULL, HY REAL DEFAULT 0 NOT NULL,
                            JA REAL DEFAULT 0 NOT NULL, JC REAL DEFAULT 0 NOT NULL,
                            JM REAL DEFAULT 0 NOT NULL, JS REAL DEFAULT 0 NOT NULL,
                            JV REAL DEFAULT 0 NOT NULL, JY REAL DEFAULT 0 NOT NULL,
                            KA REAL DEFAULT 0 NOT NULL, KC REAL DEFAULT 0 NOT NULL,
                            KM REAL DEFAULT 0 NOT NULL, KS REAL DEFAULT 0 NOT NULL,
                            KV REAL DEFAULT 0 NOT NULL, KY REAL DEFAULT 0 NOT NULL,
                            LA REAL DEFAULT 0 NOT NULL, LC REAL DEFAULT 0 NOT NULL,
                            LM REAL DEFAULT 0 NOT NULL, LS REAL DEFAULT 0 NOT NULL,
                            LV REAL DEFAULT 0 NOT NULL, LY REAL DEFAULT 0 NOT NULL,
                            MA REAL DEFAULT 0 NOT NULL, MC REAL DEFAULT 0 NOT NULL,
                            MM REAL DEFAULT 0 NOT NULL, MS REAL DEFAULT 0 NOT NULL,
                            MV REAL DEFAULT 0 NOT NULL, MY REAL DEFAULT 0 NOT NULL,
                            NA REAL DEFAULT 0 NOT NULL, NC REAL DEFAULT 0 NOT NULL,
                            NM REAL DEFAULT 0 NOT NULL, NS REAL DEFAULT 0 NOT NULL,
                            NV REAL DEFAULT 0 NOT NULL, NY REAL DEFAULT 0 NOT NULL)''')
        queryCursor.execute(Q_Create_S4_155)
        print("Done\nCreate Table S4 158...")
        Q_Create_S4_158 = ('''Create Table S4_158 (RecordID int IDENTITY(1,1) PRIMARY KEY,
                            recordDate TEXT DEFAULT 0 NOT NULL,
                            DA REAL DEFAULT 0 NOT NULL, DC REAL DEFAULT 0 NOT NULL,
                            DM REAL DEFAULT 0 NOT NULL, DS REAL DEFAULT 0 NOT NULL,
                            DV REAL DEFAULT 0 NOT NULL, DY REAL DEFAULT 0 NOT NULL,
                            EA REAL DEFAULT 0 NOT NULL, EC REAL DEFAULT 0 NOT NULL,
                            EM REAL DEFAULT 0 NOT NULL, ES REAL DEFAULT 0 NOT NULL,
                            EV REAL DEFAULT 0 NOT NULL, EY REAL DEFAULT 0 NOT NULL,
                            FA REAL DEFAULT 0 NOT NULL, FC REAL DEFAULT 0 NOT NULL,
                            FM REAL DEFAULT 0 NOT NULL, FS REAL DEFAULT 0 NOT NULL,
                            FV REAL DEFAULT 0 NOT NULL, FY REAL DEFAULT 0 NOT NULL,
                            GA REAL DEFAULT 0 NOT NULL, GC REAL DEFAULT 0 NOT NULL,
                            GM REAL DEFAULT 0 NOT NULL, GS REAL DEFAULT 0 NOT NULL,
                            GV REAL DEFAULT 0 NOT NULL, GY REAL DEFAULT 0 NOT NULL,
                            HA REAL DEFAULT 0 NOT NULL, HC REAL DEFAULT 0 NOT NULL,
                            HM REAL DEFAULT 0 NOT NULL, HS REAL DEFAULT 0 NOT NULL,
                            HV REAL DEFAULT 0 NOT NULL, HY REAL DEFAULT 0 NOT NULL,
                            JA REAL DEFAULT 0 NOT NULL, JC REAL DEFAULT 0 NOT NULL,
                            JM REAL DEFAULT 0 NOT NULL, JS REAL DEFAULT 0 NOT NULL,
                            JV REAL DEFAULT 0 NOT NULL, JY REAL DEFAULT 0 NOT NULL,
                            KA REAL DEFAULT 0 NOT NULL, KC REAL DEFAULT 0 NOT NULL,
                            KM REAL DEFAULT 0 NOT NULL, KS REAL DEFAULT 0 NOT NULL,
                            KV REAL DEFAULT 0 NOT NULL, KY REAL DEFAULT 0 NOT NULL,
                            LA REAL DEFAULT 0 NOT NULL, LC REAL DEFAULT 0 NOT NULL,
                            LM REAL DEFAULT 0 NOT NULL, LS REAL DEFAULT 0 NOT NULL,
                            LV REAL DEFAULT 0 NOT NULL, LY REAL DEFAULT 0 NOT NULL,
                            MA REAL DEFAULT 0 NOT NULL, MC REAL DEFAULT 0 NOT NULL,
                            MM REAL DEFAULT 0 NOT NULL, MS REAL DEFAULT 0 NOT NULL,
                            MV REAL DEFAULT 0 NOT NULL, MY REAL DEFAULT 0 NOT NULL,
                            NA REAL DEFAULT 0 NOT NULL, NC REAL DEFAULT 0 NOT NULL,
                            NM REAL DEFAULT 0 NOT NULL, NS REAL DEFAULT 0 NOT NULL,
                            NV REAL DEFAULT 0 NOT NULL, NY REAL DEFAULT 0 NOT NULL)''')
        queryCursor.execute(Q_Create_S4_158)
        print("Done\nCreate Table S4 160...")
        Q_Create_S4_160 = ('''Create Table S4_160 (RecordID int IDENTITY(1,1) PRIMARY KEY,
                            recordDate TEXT DEFAULT 0 NOT NULL,
                            DA REAL DEFAULT 0 NOT NULL, DC REAL DEFAULT 0 NOT NULL,
                            DM REAL DEFAULT 0 NOT NULL, DS REAL DEFAULT 0 NOT NULL,
                            DV REAL DEFAULT 0 NOT NULL, DY REAL DEFAULT 0 NOT NULL,
                            EA REAL DEFAULT 0 NOT NULL, EC REAL DEFAULT 0 NOT NULL,
                            EM REAL DEFAULT 0 NOT NULL, ES REAL DEFAULT 0 NOT NULL,
                            EV REAL DEFAULT 0 NOT NULL, EY REAL DEFAULT 0 NOT NULL,
                            FA REAL DEFAULT 0 NOT NULL, FC REAL DEFAULT 0 NOT NULL,
                            FM REAL DEFAULT 0 NOT NULL, FS REAL DEFAULT 0 NOT NULL,
                            FV REAL DEFAULT 0 NOT NULL, FY REAL DEFAULT 0 NOT NULL,
                            GA REAL DEFAULT 0 NOT NULL, GC REAL DEFAULT 0 NOT NULL,
                            GM REAL DEFAULT 0 NOT NULL, GS REAL DEFAULT 0 NOT NULL,
                            GV REAL DEFAULT 0 NOT NULL, GY REAL DEFAULT 0 NOT NULL,
                            HA REAL DEFAULT 0 NOT NULL, HC REAL DEFAULT 0 NOT NULL,
                            HM REAL DEFAULT 0 NOT NULL, HS REAL DEFAULT 0 NOT NULL,
                            HV REAL DEFAULT 0 NOT NULL, HY REAL DEFAULT 0 NOT NULL,
                            JA REAL DEFAULT 0 NOT NULL, JC REAL DEFAULT 0 NOT NULL,
                            JM REAL DEFAULT 0 NOT NULL, JS REAL DEFAULT 0 NOT NULL,
                            JV REAL DEFAULT 0 NOT NULL, JY REAL DEFAULT 0 NOT NULL,
                            KA REAL DEFAULT 0 NOT NULL, KC REAL DEFAULT 0 NOT NULL,
                            KM REAL DEFAULT 0 NOT NULL, KS REAL DEFAULT 0 NOT NULL,
                            KV REAL DEFAULT 0 NOT NULL, KY REAL DEFAULT 0 NOT NULL,
                            LA REAL DEFAULT 0 NOT NULL, LC REAL DEFAULT 0 NOT NULL,
                            LM REAL DEFAULT 0 NOT NULL, LS REAL DEFAULT 0 NOT NULL,
                            LV REAL DEFAULT 0 NOT NULL, LY REAL DEFAULT 0 NOT NULL,
                            MA REAL DEFAULT 0 NOT NULL, MC REAL DEFAULT 0 NOT NULL,
                            MM REAL DEFAULT 0 NOT NULL, MS REAL DEFAULT 0 NOT NULL,
                            MV REAL DEFAULT 0 NOT NULL, MY REAL DEFAULT 0 NOT NULL,
                            NA REAL DEFAULT 0 NOT NULL, NC REAL DEFAULT 0 NOT NULL,
                            NM REAL DEFAULT 0 NOT NULL, NS REAL DEFAULT 0 NOT NULL,
                            NV REAL DEFAULT 0 NOT NULL, NY REAL DEFAULT 0 NOT NULL)''')
        queryCursor.execute(Q_Create_S4_160)

        print("Done\nCreate Table GlobalVariables...")
        Q_Create_GlobalVariables = ('''Create Table GlobalVariables (RecordID int IDENTITY(1,1) PRIMARY KEY,
                            Variable TEXT DEFAULT 0 NOT NULL,
                            VariableDescription TEXT DEFAULT 0 NOT NULL,
                            VarValue REAL DEFAULT 0 NOT NULL)''')
        queryCursor.execute(Q_Create_GlobalVariables)

        print("Done\nPopulationg Table GlobalVariables...")
        Qry1 = (
        "INSERT INTO GlobalVariables (Variable, VariableDescription, VarValue) VALUES ('PE_S2_180', 'Plant Efficiency, S2, RD = 1.8','0.98')")
        queryCursor.execute(Qry1)
        queryCursor.commit()
        Qry2 = (
        "INSERT INTO GlobalVariables (Variable, VariableDescription, VarValue) VALUES ('PE_S2_175', 'Plant Efficiency, S2, RD = 1.75','0.98')")
        queryCursor.execute(Qry2)
        queryCursor.commit()
        Qry3 = (
        "INSERT INTO GlobalVariables (Variable, VariableDescription, VarValue) VALUES ('PE_S2_170', 'Plant Efficiency, S2, RD = 1.7','0.98')")
        queryCursor.execute(Qry3)
        queryCursor.commit()
        Qry4 = (
        "INSERT INTO GlobalVariables (Variable, VariableDescription, VarValue) VALUES ('PE_S2_165', 'Plant Efficiency, S2, RD = 1.65','0.98')")
        queryCursor.execute(Qry4)
        queryCursor.commit()
        Qry5 = (
        "INSERT INTO GlobalVariables (Variable, VariableDescription, VarValue) VALUES ('PE_S4_180', 'Plant Efficiency, S4, RD = 1.8','0.95')")
        queryCursor.execute(Qry5)
        queryCursor.commit()
        Qry6 = (
        "INSERT INTO GlobalVariables (Variable, VariableDescription, VarValue) VALUES ('PE_S4_175', 'Plant Efficiency, S4, RD = 1.75','0.95')")
        queryCursor.execute(Qry6)
        queryCursor.commit()
        Qry7 = (
        "INSERT INTO GlobalVariables (Variable, VariableDescription, VarValue) VALUES ('PE_S4_170', 'Plant Efficiency, S4, RD = 1.7','0.95')")
        queryCursor.execute(Qry7)
        queryCursor.commit()
        Qry8 = (
        "INSERT INTO GlobalVariables (Variable, VariableDescription, VarValue) VALUES ('PE_S4_165', 'Plant Efficiency, S4, RD = 1.65','0.95')")
        queryCursor.execute(Qry8)
        queryCursor.commit()
        Qry9 = (
        "INSERT INTO GlobalVariables (Variable, VariableDescription, VarValue) VALUES ('TargetCV', 'Target CV','18.95')")
        queryCursor.execute(Qry9)
        queryCursor.commit()
        Qry10 = (
        "INSERT INTO GlobalVariables (Variable, VariableDescription, VarValue) VALUES ('TargetAS', 'Target Ash','32')")
        queryCursor.execute(Qry10)
        queryCursor.commit()
        Qry11 = (
        "INSERT INTO GlobalVariables (Variable, VariableDescription, VarValue) VALUES ('BypassRatio_S2', 'Bypass Ratio for S2 (Fines)','0.35')")
        queryCursor.execute(Qry11)
        queryCursor.commit()
        Qry12 = (
        "INSERT INTO GlobalVariables (Variable, VariableDescription, VarValue) VALUES ('BypassRatio_S4', 'Bypass Ratio for S4 (Fines)','0.35')")
        queryCursor.execute(Qry12)
        queryCursor.commit()
        Qry13 = (
            "INSERT INTO GlobalVariables (Variable, VariableDescription, VarValue) VALUES ('Wash_on_Selection', 'Type of Wash selection','LessWash')")
        queryCursor.execute(Qry13)
        queryCursor.commit()

        print("Done")
        print("\nPopulating Table Dates:")
        for tablename in tables:
            #print("Table: '{}'".format(tablename), end='')
            self.populate_PySQL_tabledates(tablename)
        queryCursor.close()
        print("... Done")
        return


    def ClearPySQLData(self):
        # Clear the SQL Product Table before import new data
        print ("Clear SQL Data...")
        queryCursor = self.ConnectPySQL()
        # Q_TruncateTable = "TRUNCATE TABLE Product;"
        queryCursor.execute('DELETE FROM Product')
        queryCursor.commit()
        queryCursor.execute('VACUUM')
        queryCursor.commit()
        queryCursor.close()
        print ("Done")
        return

    def mainUi(self, Form):
        Form.setObjectName(_fromUtf8("Form"))
        # Form.setGeometry(1930, 30, 300, 1013)
        Form.setGeometry(1570, 130, 355, 800)
        Form.setWindowTitle("Main")

        frame = QtGui.QFrame()
        frame.setFrameStyle(QtGui.QFrame.Panel | QtGui.QFrame.Plain)

        # widget = QWidget(self)
        # self.setCentralWidget(widget)
        # self.overlay = Overlay(self.centralWidget())
        # self.overlay.hide()

        # self.pB_CreateSQLStructure = QtGui.QPushButton(Form)
        # self.pB_CreateSQLStructure.setObjectName(_fromUtf8("pB_CreateSQLStructure"))

        '''
        self.pB_CreatePySQLStructure = QtGui.QPushButton(Form)
        self.pB_CreatePySQLStructure.setObjectName(_fromUtf8("pB_CreatePySQLStructure"))

        self.pB_ClearSQLTable = QtGui.QPushButton(Form)
        self.pB_ClearSQLTable.setObjectName(_fromUtf8("pB_ClearSQLTable"))

        self.pB_recalc = QtGui.QPushButton(Form)
        self.pB_recalc.setObjectName(_fromUtf8("pB_recalc"))

        # self.pB_washCV = QtGui.QPushButton(Form)
        # self.pB_washCV.setObjectName(_fromUtf8("pB_washCV"))

        self.pB_washAS = QtGui.QPushButton(Form)
        self.pB_washAS.setObjectName(_fromUtf8("pB_washAS"))

        self.pB_report = QtGui.QPushButton(Form)
        self.pB_report.setObjectName(_fromUtf8("pB_report"))

        self.pB_test = QtGui.QPushButton(Form)
        self.pB_test.setObjectName(_fromUtf8("pB_test"))

        # self.pB_test2 = QtGui.QPushButton(Form)
        # self.pB_test2.setObjectName(_fromUtf8("pB_test2"))

        self.pB_targetCV = QtGui.QPushButton(Form)
        self.pB_targetCV.setObjectName((_fromUtf8("pB_targetCV")))
        self.pB_targetAS = QtGui.QPushButton(Form)
        self.pB_targetAS.setObjectName((_fromUtf8("pB_targetAS")))
        self.pB_s2_bypass = QtGui.QPushButton(Form)
        self.pB_s2_bypass.setObjectName((_fromUtf8("pB_s2_bypass")))
        self.pB_s4_bypass = QtGui.QPushButton(Form)
        self.pB_s4_bypass.setObjectName((_fromUtf8("pB_s4_bypass")))

        self.pB_s4_rd_180_pe = QtGui.QPushButton(Form)
        self.pB_s4_rd_180_pe.setObjectName((_fromUtf8("pB_s4_rd_180_pe")))
        self.pB_s4_rd_175_pe = QtGui.QPushButton(Form)
        self.pB_s4_rd_175_pe.setObjectName((_fromUtf8("pB_s4_rd_175_pe")))
        self.pB_s4_rd_170_pe = QtGui.QPushButton(Form)
        self.pB_s4_rd_170_pe.setObjectName((_fromUtf8("pB_s4_rd_170_pe")))
        self.pB_s4_rd_165_pe = QtGui.QPushButton(Form)
        self.pB_s4_rd_165_pe.setObjectName((_fromUtf8("pB_s4_rd_165_pe")))

        self.pB_s2_rd_180_pe = QtGui.QPushButton(Form)
        self.pB_s2_rd_180_pe.setObjectName((_fromUtf8("pB_s2_rd_180_pe")))
        self.pB_s2_rd_175_pe = QtGui.QPushButton(Form)
        self.pB_s2_rd_175_pe.setObjectName((_fromUtf8("pB_s2_rd_175_pe")))
        self.pB_s2_rd_170_pe = QtGui.QPushButton(Form)
        self.pB_s2_rd_170_pe.setObjectName((_fromUtf8("pB_s2_rd_170_pe")))
        self.pB_s2_rd_165_pe = QtGui.QPushButton(Form)
        self.pB_s2_rd_165_pe.setObjectName((_fromUtf8("pB_s2_rd_165_pe")))

        self.s2_info_field = QtGui.QLabel(Form)
        self.s2_info_field.setObjectName(_fromUtf8("s2_info_field"))

        self.s4_info_field = QtGui.QLabel(Form)
        self.s4_info_field.setObjectName(_fromUtf8("s4_info_field"))

        self.pe_180_info_field = QtGui.QLabel(Form)
        self.pe_180_info_field.setObjectName(_fromUtf8("pe_180_info_field"))
        self.pe_175_info_field = QtGui.QLabel(Form)
        self.pe_175_info_field.setObjectName(_fromUtf8("pe_175_info_field"))
        self.pe_170_info_field = QtGui.QLabel(Form)
        self.pe_170_info_field.setObjectName(_fromUtf8("pe_170_info_field"))
        self.pe_165_info_field = QtGui.QLabel(Form)
        self.pe_165_info_field.setObjectName(_fromUtf8("pe_165_info_field"))

        self.pB_test2 = QtGui.QPushButton(Form)
        self.pB_test2.setObjectName(_fromUtf8("pB_test2"))

        # ROM AD CV Before Wash
        self.pB_plot_CV = QtGui.QPushButton(Form)
        self.pB_plot_CV.setObjectName(_fromUtf8("pB_plot_CV"))

        self.pB_plot_CV_after_wash = QtGui.QPushButton(Form)
        self.pB_plot_CV_after_wash.setObjectName(_fromUtf8("pB_plot_CV_after_wash"))

        self.pB_plot_tons_washed = QtGui.QPushButton(Form)
        self.pB_plot_tons_washed.setObjectName(_fromUtf8("pB_plot_tons_washed"))

        self.pB_wash_selection = QtGui.QPushButton(Form)
        self.pB_wash_selection.setObjectName(_fromUtf8("pB_wash_selection"))
        # pe_180_info_field

        self.status_field = QtGui.QLabel(Form)
        self.status_field.setObjectName(_fromUtf8("status_field"))
        self.status_date_field = QtGui.QLabel(Form)
        self.status_date_field.setObjectName(_fromUtf8("status_date_field"))'''

        # update_TargetCV_value(self, varValue):

        self.retranslateUi(Form)
        QtCore.QMetaObject.connectSlotsByName(Form)

    # Mapping of signals for the Main User Interface:
    def retranslateUi(self, Form):

        # self.pB_CreateSQLStructure.setText(_translate("Form", "Create SQL Structure", None))
        # self.pB_CreateSQLStructure.setGeometry(QtCore.QRect(25, 110, 250, 23))
        # self.pB_CreateSQLStructure.clicked.connect(self.CreateSQLStructure)

       '''
        try:
            self.pB_CreatePySQLStructure.setText(_translate("Form", "Create PySQL Structure", None))
            self.pB_CreatePySQLStructure.setGeometry(QtCore.QRect(25, 110, 200, 23))
            self.pB_CreatePySQLStructure.clicked.connect(self.CreatePySQL)
            self.pB_CreatePySQLStructure.setEnabled(False)
        except:
            pass
        try:
            self.pB_ClearSQLTable.setText(_translate("Form", "Clear SQL Table", None))
            self.pB_ClearSQLTable.setGeometry(QtCore.QRect(25, 140, 200, 23))
            self.pB_ClearSQLTable.clicked.connect(self.ClearSQLData)
            # self.pB_ClearSQLTable.clicked.connect(self.ClearPySQLData)
            self.pB_ClearSQLTable.setEnabled(False)
        except:
            pass

        try:
            self.pB_recalc.setText(_translate("Form", "Re-calculate to AD Values", None))
            self.pB_recalc.setGeometry(QtCore.QRect(25, 170, 200, 23))
            self.pB_recalc.clicked.connect(self.recalc)
        except:
            pass

        try:
            self.pB_washAS.setText(_translate("Form", "Wash on AS", None))
            self.pB_washAS.setGeometry(QtCore.QRect(25, 260, 200, 23))
            self.pB_washAS.setEnabled(False)
        except:
            pass

        try:
            self.pB_test.setText(_translate("Form", "Wash on CV, least wash tons", None))
            self.pB_test.setGeometry(QtCore.QRect(25, 200, 200, 23))
            self.pB_test.clicked.connect(self.simulation150)
        except:
            pass



        try:
            self.pB_test2.setText(_translate("Form", "pB_test2", None))
            self.pB_test2.setGeometry(QtCore.QRect(25, 700, 250, 23))
            self.pB_test2.clicked.connect(self.checkinifile)

            self.pB_report.setText(_translate("Form", "Report", None))
            self.pB_report.setGeometry(QtCore.QRect(25, 290, 200, 23))
            self.pB_report.clicked.connect(self.report)
        except:
            pass

        try:
            self.pB_plot_CV.setText(_translate("Form", "pB_plot_CV", None))
            self.pB_plot_CV.setGeometry(QtCore.QRect(180, 610, 150, 23))
            self.pB_plot_CV.clicked.connect(self.plot_CV)
        except:
            pass



        try:

            self.pB_plot_CV_after_wash.setText(_translate("Form", "CV after wash", None))
            self.pB_plot_CV_after_wash.setGeometry(QtCore.QRect(180, 640, 150, 23))
            self.pB_plot_CV_after_wash.clicked.connect(self.plot_CV_after_wash)

            self.pB_plot_tons_washed.setText(_translate("Form", "TON washed", None))
            self.pB_plot_tons_washed.setGeometry(QtCore.QRect(180, 670, 150, 23))
            self.pB_plot_tons_washed.clicked.connect(self.plot_tons_washed)


            # self.statusField.setGeometry(QtCore.QRect(25, 600, 250, 23))
            # self.statusField.setText(_translate("Form", "status", None))
            # self.statusField.setVisible(False)

            targetCV, variableDesc = self.getGlobalVariable('TargetCV')
            targetCV_text = ("{}".format(targetCV))
            self.pB_targetCV.setText(_translate("Form", ("Target CV: {}".format(targetCV_text)), None))
            self.pB_targetCV.setGeometry(QtCore.QRect(230, 200, 100, 23))
            self.pB_targetCV.clicked.connect(self.adjustTargetCV)

            wash_selection, variableDesc = self.getGlobalVariable('Wash_on_Selection')
            wash_selection_text = ("{}".format(wash_selection))
            self.pB_wash_selection.setText(_translate("Form", ("Wash: {}".format(wash_selection_text)), None))
            self.pB_wash_selection.setGeometry(QtCore.QRect(230, 230, 100, 23))
            # self.pB_wash_selection.clicked.connect(self.adjustTargetCV)

        except:
            pass

        try:
            targetAS, variableDesc = self.getGlobalVariable('TargetAS')
            targetAS_text = ("{}".format(targetAS))
            self.pB_targetAS.setText(_translate("Form", ("Target AS: {}".format(targetAS_text)), None))
            self.pB_targetAS.setGeometry(QtCore.QRect(230, 260, 100, 23))
            self.pB_targetAS.clicked.connect(self.adjustTargetAS)
        except:
            pass

        try:
            s2_bypass, variableDesc = self.getGlobalVariable('BypassRatio_S2')
            s2_bypass_text = ("{}".format(s2_bypass))
            self.pB_s2_bypass.setText(_translate("Form", ("BypassRatio S2: {}".format(s2_bypass_text)), None))
            self.pB_s2_bypass.setGeometry(QtCore.QRect(25, 350, 150, 23))
            self.pB_s2_bypass.clicked.connect(self.adjust_s2_bypassfactor)
        except:
            pass

        try:
            s4_bypass, variableDesc = self.getGlobalVariable('BypassRatio_S4')
            s4_bypass_text = ("{}".format(s4_bypass))
            self.pB_s4_bypass.setText(_translate("Form", ("BypassRatio S4: {}".format(s4_bypass_text)), None))
            self.pB_s4_bypass.setGeometry(QtCore.QRect(180, 350, 150, 23))
            self.pB_s4_bypass.clicked.connect(self.adjust_s4_bypassfactor)
        except:
            pass

        try:
            s4_rd_180_pe, variableDesc = self.getGlobalVariable('PE_S4_180')
            s4_rd_180_pe_text = ("{}".format(s4_rd_180_pe))
            self.pB_s4_rd_180_pe.setText(_translate("Form", ("{}".format(s4_rd_180_pe_text)), None))
            self.pB_s4_rd_180_pe.setGeometry(QtCore.QRect(175, 410, 50, 23))
            self.pB_s4_rd_180_pe.clicked.connect(self.adjust_s4_rd_180_pe)

            s4_rd_175_pe, variableDesc = self.getGlobalVariable('PE_S4_175')
            s4_rd_175_pe_text = ("{}".format(s4_rd_175_pe))
            self.pB_s4_rd_175_pe.setText(_translate("Form", ("{}".format(s4_rd_175_pe_text)), None))
            self.pB_s4_rd_175_pe.setGeometry(QtCore.QRect(175, 440, 50, 23))
            self.pB_s4_rd_175_pe.clicked.connect(self.adjust_s4_rd_175_pe)

            s4_rd_170_pe, variableDesc = self.getGlobalVariable('PE_S4_170')
            s4_rd_170_pe_text = ("{}".format(s4_rd_170_pe))
            self.pB_s4_rd_170_pe.setText(_translate("Form", ("{}".format(s4_rd_170_pe_text)), None))
            self.pB_s4_rd_170_pe.setGeometry(QtCore.QRect(175, 470, 50, 23))
            self.pB_s4_rd_170_pe.clicked.connect(self.adjust_s4_rd_170_pe)

            s4_rd_165_pe, variableDesc = self.getGlobalVariable('PE_S4_165')
            s4_rd_165_pe_text = ("{}".format(s4_rd_165_pe))
            self.pB_s4_rd_165_pe.setText(_translate("Form", ("{}".format(s4_rd_165_pe_text)), None))
            self.pB_s4_rd_165_pe.setGeometry(QtCore.QRect(175, 500, 50, 23))
            self.pB_s4_rd_165_pe.clicked.connect(self.adjust_s4_rd_165_pe)

            s2_rd_180_pe, variableDesc = self.getGlobalVariable('PE_S2_180')
            s2_rd_180_pe_text = ("{}".format(s2_rd_180_pe))
            self.pB_s2_rd_180_pe.setText(_translate("Form", ("{}".format(s2_rd_180_pe_text)), None))
            self.pB_s2_rd_180_pe.setGeometry(QtCore.QRect(120, 410, 50, 23))
            self.pB_s2_rd_180_pe.clicked.connect(self.adjust_s2_rd_180_pe)

            s2_rd_175_pe, variableDesc = self.getGlobalVariable('PE_S2_175')
            s2_rd_175_pe_text = ("{}".format(s2_rd_175_pe))
            self.pB_s2_rd_175_pe.setText(_translate("Form", ("{}".format(s2_rd_175_pe_text)), None))
            self.pB_s2_rd_175_pe.setGeometry(QtCore.QRect(120, 440, 50, 23))
            self.pB_s2_rd_175_pe.clicked.connect(self.adjust_s2_rd_175_pe)

            s2_rd_170_pe, variableDesc = self.getGlobalVariable('PE_S2_170')
            s2_rd_170_pe_text = ("{}".format(s2_rd_170_pe))
            self.pB_s2_rd_170_pe.setText(_translate("Form", ("{}".format(s2_rd_170_pe_text)), None))
            self.pB_s2_rd_170_pe.setGeometry(QtCore.QRect(120, 470, 50, 23))
            self.pB_s2_rd_170_pe.clicked.connect(self.adjust_s2_rd_170_pe)

            s2_rd_165_pe, variableDesc = self.getGlobalVariable('PE_S2_165')
            s2_rd_165_pe_text = ("{}".format(s2_rd_165_pe))
            self.pB_s2_rd_165_pe.setText(_translate("Form", ("{}".format(s2_rd_165_pe_text)), None))
            self.pB_s2_rd_165_pe.setGeometry(QtCore.QRect(120, 500, 50, 23))
            self.pB_s2_rd_165_pe.clicked.connect(self.adjust_s2_rd_165_pe)
        except:
            pass

        self.s2_info_field.setGeometry(QtCore.QRect(130, 380, 50, 23))
        self.s2_info_field.setText(_translate("Form", "S2", None))
        self.s2_info_field.setVisible(True)

        self.s4_info_field.setGeometry(QtCore.QRect(180, 380, 50, 23))
        self.s4_info_field.setText(_translate("Form", "S4", None))
        self.s4_info_field.setVisible(True)

        self.pe_180_info_field.setGeometry(QtCore.QRect(25, 410, 120, 23))
        self.pe_180_info_field.setText(_translate("Form", "P/E @ RD 1.80", None))
        self.pe_180_info_field.setVisible(True)

        self.pe_175_info_field.setGeometry(QtCore.QRect(25, 440, 120, 23))
        self.pe_175_info_field.setText(_translate("Form", "P/E @ RD 1.75", None))
        self.pe_175_info_field.setVisible(True)

        self.pe_170_info_field.setGeometry(QtCore.QRect(25, 470, 120, 23))
        self.pe_170_info_field.setText(_translate("Form", "P/E @ RD 1.70", None))
        self.pe_170_info_field.setVisible(True)

        self.pe_165_info_field.setGeometry(QtCore.QRect(25, 500, 120, 23))
        self.pe_165_info_field.setText(_translate("Form", "P/E @ RD 1.65", None))
        self.pe_165_info_field.setVisible(True)

        self.status_field.setGeometry(QtCore.QRect(5, 5, 345, 23))
        self.status_field.setText(_translate("Form", "", None))
        # self.status_field.setVisible(False)


        self.status_date_field.setGeometry(QtCore.QRect(250, 35, 100, 23))
        self.status_date_field.setText(_translate("Form", "", None))
        # self.status_date_field.setVisible(False)
        :param Form:
        :return:
        '''


        # self.status_field.setText(_translate("Form", "", None))
        # self.status_field.setVisible(False)
        # self.status_date_field.setText(_translate("Form", "", None))
        # self.status_date_field.setVisible(False)

        # self.pe_180_info_field.setGeometry(QtCore.QRect(25, 290, 150, 23))
        # self.pe_180_info_field.setText(_translate("Form", "Plant Efficiency @ RD 1.80", None))
        # self.pe_180_info_field.setVisible(True)

        # self.pB_washCV.setText(_translate("Form", "Wash on CV", None))
        # self.pB_washCV.setGeometry(QtCore.QRect(25, 230, 250, 23))
        # self.pB_washCV.clicked.connect(self.washSequence)
        # TODO: Fix this
        # self.pB_washCV.clicked.connect(self.washON('CV'))
        # self.pB_washCV.setFlat()

        # TODO: Fix this
        # self.pB_washAS.clicked.connect(self.washON('AS'))
        # self.pB_washAS.setFlat()

        # self.pB_report.setText(_translate("Form", "Wash on CV & Report", None))
        # self.pB_report.setGeometry(QtCore.QRect(25, 290, 250, 23))
        # self.pB_report.clicked.connect(self.report)

        # self.pB_wash.setText(_translate("Form", "Wash Simulation on CV", None))
        # self.pB_wash.setGeometry(QtCore.QRect(25, 200, 250, 23))
        # self.pB_wash.clicked.connect(self.washSequence)

        # Main Program:

if __name__ == "__main__":
    try:
        # spin = Overlay()
        app = QtGui.QApplication(sys.argv)
        myapp = MainWindow()
        myapp.show()
        sys.exit(app.exec_())
    except KeyboardInterrupt:
        print ("\n\Program aborted by user...\n")
        # spin.stop()
        sys.exit()


